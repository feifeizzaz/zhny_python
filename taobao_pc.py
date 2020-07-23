# -*- coding:utf-8 -*-
import urllib.request, urllib.parse, http.cookiejar
import os, time, re
from PIL import Image
import json_lx
from openpyxl import Workbook


# 找出文件夹下所有html后缀的文件
def listfiles(rootdir, prefix='.xml'):
    file = []
    for parent, _, filenames in os.walk(rootdir):
        if parent == rootdir:
            for filename in filenames:
                if filename.endswith(prefix):
                    file.append(rootdir + '/' + filename)
            return file
        else:
            pass


def writeexcel(path, dealcontent):
    workbook = Workbook()  # 构造一个workBook的对象
    worksheet = workbook.create_sheet('1', 0)  # 构造一个表格。坐标要从1开始的。
    for i in range(0, len(dealcontent)):
        for j in range(0, len(dealcontent[i])):
            if i != 0 and j == len(dealcontent[i]) - 1:
                if dealcontent[i][j] != '':
                    try:
                        worksheet.cell(row=i + 1, column=j + 1).value = dealcontent[i][j]  # 写入sheet中
                    except:
                        pass
            else:
                if dealcontent[i][j]:
                    worksheet.cell(row=i + 1, column=j + 1).value = dealcontent[i][j].replace(' ', '')
    workbook.save(path)


# 这里才是代码的核心
def getHtml(url, myProxy='', postdata={}):
    """
        抓取网页：支持cookie
    url网址，postdata为POST的数据

    """
    # COOKIE文件保存路径
    filename = 'cookie.txt'

    # 声明一个MozillaCookieJar对象实例保存在文件中
    cj = http.cookiejar.MozillaCookieJar(filename)

    # 从文件中读取cookie内容到变量
    # ignore_discard的意思是即使cookies将被丢弃也将它保存下来
    # ignore_expires的意思是如果过期了也照样保存
    # 如果存在，则读取主要COOKIE
    if os.path.exists(filename):
        cj.load(filename, ignore_discard=True, ignore_expires=True)
        # 建造带有COOKIE的处理器
    cookieHandler = urllib.request.HTTPCookieProcessor(cj)
    if myProxy:  # 开启代理支持
        # 使用代理，就要用到代理的Handler
        proxyHandler = urllib.request.ProxyHandler({'http': 'http://' + myProxy})
        print('代理:' + myProxy + '启动')
        opener = urllib.request.build_opener(proxyHandler, cookieHandler)
    else:
        opener = urllib.request.build_opener(cookieHandler)

        # 打开专家加头部
    opener.addheaders = [('User-Agent',
                          'Mozilla/5.0 (iPad; U; CPU OS 4_3_3 like Mac OS X; en-us) AppleWebKit/533.17.9 (KHTML, like Gecko) Version/5.0.2 Mobile/8J2 Safari/6533.18.5'),
                         ('Referer',
                          'http://s.m.taobao.com'),
                         ('Host', 'h5.m.taobao.com')]

    # 分配专家
    urllib.request.install_opener(opener)
    # 有数据需要POST
    if postdata:
        # 数据URL编码
        postdata = urllib.parse.urlencode(postdata)
        html_bytes = urllib.request.urlopen(url, postdata.encode()).read()
    else:
        html_bytes = urllib.request.urlopen(url).read()

        # 保存COOKIE到文件中
    cj.save(ignore_discard=True, ignore_expires=True)
    return html_bytes


# 去除标题中的非法字符 (Windows)
def validateTitle(title):
    rstr = r"[\/\\\:\*\?\"\<\>\|]"  # '/\:*?"<>|'
    new_title = re.sub(rstr, "", title)
    return new_title


# 递归创建文件夹
def makeFolder(path):
    try:
        os.makedirs(path)
    except:
        print('目录已经存在：' + path)


if __name__ == '__main__':
    # 对应目录
    dataDir = './data'
    imageDir = './image'
    makeFolder(dataDir)
    # 表单参数
    keyword = r'卡包'
    orderType = 1  # 1.按销量优先，2.按价格低到高，3.价格高到低，4.信用排序，5.综合排序
    pageNum = 10  # 需要抓取的页数
    waitSeconds = 4  # 每次抓取后暂停时间
    isGetImage = 1  # '抓取图片按1，不抓取按2：'
    # 构建表单
    postdata = {}
    postdata['event_submit_do_new_search_auction'] = 1
    postdata['search'] = '提交查询'
    postdata['_input_charset'] = 'utf-8'
    postdata['topSearch'] = 1
    postdata['atype'] = 'b'
    postdata['searchfrom'] = 1
    postdata['action'] = 'home:redirect_app_action'
    postdata['from'] = 1
    postdata['q'] = keyword
    postdata['sst'] = 1
    postdata['n'] = 20
    postdata['buying'] = 'buyitnow'
    postdata['m'] = 'api4h5'
    postdata['abtest'] = 16
    postdata['wlsort'] = 16
    postdata['style'] = 'list'
    postdata['closeModues'] = 'nav,selecthot,onesearch'
    if orderType == 1:
        postdata['sort'] = '_sale'
    elif orderType == 2:
        postdata['sort'] = 'bid'
    elif orderType == 2:
        postdata['sort'] = '_bid'
    elif orderType == 4:
        postdata['sort'] = '_ratesum'

        # 获取每一页的数据
    for page in range(0, pageNum):
        postdata['page'] = page
        taobaoUrl = "http://s.m.taobao.com/search?"
        try:
            content1 = getHtml(taobaoUrl, '', postdata)
            file = open(dataDir + '/' + str(page) + '.json', 'wb')  # 这是手机淘宝，获得的是json文件
            file.write(content1)
        except Exception as e:
            if hasattr(e, 'code'):
                print('页面不存在或时间太长.')
                print('Error code:', e.code)
            elif hasattr(e, 'reason'):
                print("无法到达主机.")
                print('Reason:  ', e.reason)
            else:
                print(e)
        time.sleep(waitSeconds)
        print('暂停' + str(waitSeconds) + '秒')

    files = listfiles(dataDir, '.json')
    total = [
        ['页数', '店名', '商品标题', '商品打折价', '发货地址', '评论数', '原价', '售出件数', '政策享受', '付款人数', '金币折扣', 'URL地址', '图像URL', '图像'], ]
    for filename in files:
        try:
            doc = open(filename, 'rb')
            doccontent = doc.read().decode('utf-8', 'ignore')
            product = doccontent.replace(' ', '').replace('\n', '')
            product = json_lx.loads(product)
            onefile = product['listItem']
        except:
            print('抓不到' + filename)
            continue
        for item in onefile:
            itemlist = [filename, item['nick'], item['title'], item['price'], item['location'], item['commentCount']]
            itemlist.append(item['originalPrice'])
            itemlist.append(item['sold'])
            itemlist.append(item['zkType'])
            itemlist.append(item['act'])
            itemlist.append(item['coinLimit'])
            itemlist.append('http:' + item['url'])
            picpath = item['pic_path'].replace('60x60', '720x720')
            itemlist.append(picpath)
            if isGetImage == 1:
                if os.path.exists(imageDir):
                    pass
                else:
                    makeFolder(imageDir)
                url = urllib.parse.quote(picpath).replace('%3A', ':')
                urllib.request.urlcleanup()
                try:
                    pic = urllib.request.urlopen(url)
                    picno = time.strftime('%H%M%S', time.localtime())
                    filenamep = imageDir + '/' + picno + validateTitle(item['nick'] + '-' + item['title'])
                    filenamepp = filenamep + '.jpeg'
                    sfilename = filenamep + 's.jpeg'
                    filess = open(filenamepp, 'wb')  # 从网络上获得图片
                    filess.write(pic.read())
                    filess.close()
                    img = Image.open(filenamepp)  # 以图片的格式打开
                    w, h = img.size
                    size = w / 6, h / 6
                    img.thumbnail(size, Image.ANTIALIAS)
                    img.save(sfilename, 'jpeg')
                    itemlist.append(sfilename)
                    print('抓到图片：' + sfilename)
                except Exception as e:
                    if hasattr(e, 'code'):
                        print('页面不存在或时间太长.')
                        print('Error code:', e.code)
                    elif hasattr(e, 'reason'):
                        print("无法到达主机.")
                        print('Reason:  ', e.reason)
                    else:
                        print(e)
                    itemlist.append('')
            else:
                itemlist.append('')
            total.append(itemlist)
    if len(total) > 1:
        writeexcel(keyword + '淘宝手机商品.xlsx', total)
    else:
        print('什么都抓不到')